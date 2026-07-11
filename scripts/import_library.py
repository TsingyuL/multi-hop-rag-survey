#!/usr/bin/env python3
"""Convert the private working literature workbook into a public, auditable import.

The imported file preserves the workbook's original folders as `library_categories`.
It is intentionally separate from methods.csv: importing a record does not verify its
metadata, establish scope, or assign the survey taxonomy.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = ROOT / "survey" / "MH_QA_Library.xlsx"
DEFAULT_OUTPUT = ROOT / "taxonomy" / "library_papers.csv"
FIELDS = [
    "library_id", "title", "authors", "year", "venue", "arxiv_id", "source_url",
    "library_categories", "source_origin", "priority", "notes", "review_state",
]


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalized(value: str) -> str:
    value = unicodedata.normalize("NFKD", value).lower()
    return re.sub(r"[^a-z0-9]+", "", value)


def find_header(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int]] | None:
    for index, row in enumerate(rows[:6]):
        headers = {text(value): column for column, value in enumerate(row) if text(value)}
        if "标题" in headers and "年份" in headers:
            return index, headers
    return None


def first(record: dict[str, str], field: str, value: str) -> None:
    if not record[field] and value:
        record[field] = value


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    records: dict[str, dict[str, str]] = {}
    categories: defaultdict[str, set[str]] = defaultdict(set)
    sources: defaultdict[str, set[str]] = defaultdict(set)
    triage: defaultdict[str, bool] = defaultdict(bool)
    title_keys: dict[str, str] = {}

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        found = find_header(rows)
        if not found:
            continue
        header_row, headers = found
        for row in rows[header_row + 1:]:
            def get(name: str) -> str:
                position = headers.get(name)
                return text(row[position]) if position is not None and position < len(row) else ""

            title = get("标题")
            if not title:
                continue
            arxiv_id = get("arXiv ID")
            source_url = get("arXiv 链接")
            if not source_url and arxiv_id:
                source_url = f"https://arxiv.org/abs/{arxiv_id}"
            normalized_title = normalized(title)
            arxiv_key = f"arxiv:{arxiv_id.lower()}" if arxiv_id else ""
            # The workbook's summary sheet often omits arXiv IDs while detailed
            # category sheets provide them. Title matching bridges those views.
            key = arxiv_key if arxiv_key in records else title_keys.get(normalized_title, arxiv_key or f"title:{normalized_title}")
            if key not in records:
                records[key] = {
                    "library_id": "lib_" + hashlib.sha1(key.encode("utf-8")).hexdigest()[:12],
                    "title": title,
                    "authors": get("作者"),
                    "year": get("年份"),
                    "venue": get("会议/期刊"),
                    "arxiv_id": arxiv_id,
                    "source_url": source_url,
                    "library_categories": "",
                    "source_origin": "",
                    "priority": get("优先级"),
                    "notes": get("备注"),
                    "review_state": "imported",
                }
            title_keys[normalized_title] = key
            record = records[key]
            for field, value in (("authors", get("作者")), ("year", get("年份")), ("venue", get("会议/期刊")),
                                 ("arxiv_id", arxiv_id), ("source_url", source_url), ("priority", get("优先级")),
                                 ("notes", get("备注"))):
                first(record, field, value)
            category = get("子类") or worksheet.title
            origin = get("来源") or worksheet.title
            categories[key].add(category)
            sources[key].add(origin)
            triage[key] = triage[key] or "建议清理" in origin or "误归类" in title

    for key, record in records.items():
        record["library_categories"] = ";".join(sorted(categories[key]))
        record["source_origin"] = ";".join(sorted(sources[key]))
        record["review_state"] = "needs_triage" if triage[key] else "imported"

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(sorted(records.values(), key=lambda row: (row["year"] or "0000", row["title"].casefold()), reverse=True))
    print(f"Imported {len(records)} unique records from {args.input.relative_to(ROOT)} into {args.output.relative_to(ROOT)}.")


if __name__ == "__main__":
    main()
